import openpyxl
import socket
import serial
import time
import visa
import os
from time import sleep, gmtime, strftime
from openpyxl.styles import PatternFill

#**********************User Setup***********************
GEN_IP = "169.254.181.61"       #GEN power supply using LAN
GEN_COM = "COM4"                 #GEN power supply using RS
GEN_BAUD = 19200                  #GEN baud rate
GEN_USB = "USB0::0x0B21::0x0025::43325545313630313856::INSTR" #GEN power supply using USB
GEN_IEEE = "GPIB::6::INSTR"     #GEN power supply using GPIB
DMM1_IP = "xxx.xxx.xx.xx"
DMM2_IP = "169.254.5.55"
DMM3_IP = "169.254.1.1"
LOAD_IP = "169.254.72.43"
SCOPE_USB = "USB0::0x0B21::0x003C::91L423897-1::1::INSTR"
DPM_USB = "USB0::0x0B21::0x0025::43325545313630313856::INSTR"
DFLT_DELAY = "0.1"
DFLT_PROTOCOL = "RS-232"               #TCP, UDP, VISA, RS-232 for GEN PS
LONG_TERM_LOG = "1"                 #Save log each iteration, On=1, Off=0
LOOP_AMOUNT = "3"                   #Loop "x" times (inf = infinite)
EXCEL_SOURCE = "Inductive_Load_Test(-25 - 25)_v01.xlsx"
EXCEL_RESULTS = "Inductive_Load_Test(-25 - 25)_v01_RESULTS.xlsx"
#*******************************************************

#Setup
date = strftime("%a,%d %b %Y %I:%M:%S %p")
fName = os.path.basename(__file__)
TCP_PORT = 8003
UDP_PORT = 8005
BUFFER_SIZE = 800
#rm=visa.ResourceManager()
#rm.list_resources()

print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
print "     "
IP = raw_input("Press Enter to Start Test...")
print "           "
print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
print "     "

# Setup Delay
delay = raw_input("Please enter the delay in between commands in (sec), <Enter> for " + DFLT_DELAY + ": ")
if len(delay) < 1:
    delay = DFLT_DELAY
delay = float(delay)
print "    OK."
print "    Delay: " + str(delay)
print "     "


# Choose PS Protocol:
if DFLT_PROTOCOL == "TCP":
    prot = 1
if DFLT_PROTOCOL == "UDP":
    prot = 2
if DFLT_PROTOCOL == "VISA":
    prot = 3
if DFLT_PROTOCOL == "RS-232":
    prot = 4
print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"


# Select workbook
wb = openpyxl.load_workbook(EXCEL_SOURCE) #PUT NAME OF EXCEL SOURCE HERE
sheet = wb.active
print sheet
print "     "
MaxRow = sheet.max_row
rowNum = 5

#Variables
ind = 1         #Previous comment indicator
dataC = 0       #Data Compare
expectC = 0     #Expect Compare
errCount = 0    #Error Counter
comCount = 0    #Command Counter
loopCount = 0   #Loop Coutner
timCount = 0    #Time-out Counter
timFlag = 0     #Time-out Flag
e = 0           #Error Opening Instrument 

# Print at top of workbook
sheet.cell(row=2,column=1).value = date                             #Date
if prot == 4:
    sheet.cell(row=2,column=2).value = "GEN COM: " +GEN_COM           #Serial Port
    sheet.cell(row=2,column=4).value = "BAUDRATE:" +str(GEN_BAUD)
else:
    sheet.cell(row=2,column=2).value = "GEN IP Address: " +GEN_IP           #IP Address
sheet.cell(row=2,column=3).value = "Delay in (sec): " + str(delay)  #Delay
if prot != 4:
    sheet.cell(row=2,column=4).value = "Protocol: " + DFLT_PROTOCOL     #Protocol
sheet.cell(row=1,column=1).value = "Workbook: " + EXCEL_SOURCE      #Workbook
sheet.cell(row=1,column=2).value = "Script: " + fName               #Script




# Main EXCEL: Commands, Queries, other Types
while rowNum < MaxRow:  # Go from rowNum - last row

  #*** GEN Power Supply: LAN ****************************************************
           
  #Command PS LAN       
  if sheet.cell(row=rowNum,column=1).value == 'PSL Command':
   if ind == 1:
       print "           "
       print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"    
   MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
   print rowNum,
   ind = 0
   if prot == 1:
     s1.send(MESSAGE)
   if prot == 2:
     s1.sendto(MESSAGE, (IP, UDP_PORT))
   if prot == 3:
     s1.write(MESSAGE)
   print " Send Command: " + MESSAGE
   #print "\n"
   comCount += 1
   sleep(delay)
   

  #Fast Command PS LAN
  if sheet.cell(row=rowNum,column=1).value == 'PSL CommandF':
   if ind == 1:
       print "           "
       print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
   MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
   print rowNum,
   ind = 0
   if prot == 1:
     s1.send(MESSAGE)
   if prot == 2:
     s1.sendto(MESSAGE, (IP, UDP_PORT))
   if prot == 3:
     s1.write(MESSAGE)
   print " Send Command: " + MESSAGE
   #print "\n"
   comCount += 1
   sleep(0)
           
  #Query PS LAN     
  if sheet.cell(row=rowNum,column=1).value == 'PSL Query':
   if ind == 1:
       print "           "
       print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
   MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
   print rowNum,
   ind = 0
   print " Send Query:   " + MESSAGE
   if prot == 1:
     s1.send(MESSAGE)
     try:
       data = s1.recv(BUFFER_SIZE).strip()
     except socket.timeout:
       data = "ERR: Time-out"
   if prot == 2:
     s1.sendto(MESSAGE, (IP, UDP_PORT))
     try:
       data, addr = s1.recvfrom(BUFFER_SIZE)
       data  = data.strip()
     except socket.timeout:
       data = "ERR: Time-out"
   if prot == 3:
     try:
       data = s1.query(MESSAGE).strip()
     except visa.VisaIOError:
       data = "ERR: Time-out"
   sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
   dataC = data 
   if dataC == "ERR: Time-out":
       timFlag = 1
       timCount += 1
   expectC = sheet.cell(row=rowNum,column=4).value
   print "     Expected:  " + str(expectC).rstrip()
   print "     Actual:    " + dataC.rstrip()
   #print "\n"
   comCount += 1
   sleep(delay)

  #Open Socket PS LAN
  if sheet.cell(row=rowNum,column=1).value == 'PSL Open':
      if prot == 1:
          s1 = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
          s1.connect((IP, TCP_PORT))
          s1.settimeout(3)
          PrntProt = "    Open Socket: TCP"
      if prot == 2:
          s1 = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
          s1.settimeout(3)
          PrntProt = "    Open Socket: UDP"
      if prot == 3:
          s1 = rm.open_resource("TCPIP0::" + IP + "::inst0::INSTR")
          rm = visa.ResourceManager()
          s1.timeout = 3000
          PrntProt = "    Open Socket: VISA"
      print PrntProt

  #Close Socket PS LAN
  if sheet.cell(row=rowNum,column=1).value == 'PSL Close':
      s1.close()


  #*** GEN Power Supply: RS-232 ****************************************************
  #Open PS Serial
  if sheet.cell(row=rowNum,column=1).value == 'PSR Open':
      ser = serial.Serial(port = GEN_COM, baudrate=GEN_BAUD,
                           bytesize=8, timeout=2, stopbits=serial.STOPBITS_ONE)
      try:
          #ser.open()
          print "*** GEN Serial Connected ***"
      except serial.SerialException:
          print "Error: Serial already open or not found"
          
          
  #Command PS Serial
  if sheet.cell(row=rowNum,column=1).value == 'PSR Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value
      ser.write(MESSAGE.encode() + '\r\n')
      print " Write: " + MESSAGE
      sleep(delay)
      if ser.inWaiting() > 0:
          data = ser.readline()
      else :
          print "Error"
      print " Read: " + data
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
  #Close PS Serial
  if sheet.cell(row=rowNum,column=1).value == 'PSR Close':
      print rowNum,
      print "*** GEN Serial Disconnected ***"
      ser.close()
   

    #*** Power Supply GEN: USB ****************************************************************
  #Open PSUSB
  if sheet.cell(row=rowNum,column=1).value == 'PSU Open':
      try:
        P1 = rm.open_resource(GEN_USB)
      except visa.VisaIOError:
        print "*** GEN USB Not Found ***"
        e = 1
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      if e != 1:
        print "*** GEN USB Connected ***"
        ind = 0

  #Command PSUSB
  if sheet.cell(row=rowNum,column=1).value == 'PSU Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      P1.write(MESSAGE)
      print " Send PS Command: " + MESSAGE
      sleep(delay)
      
  #Query PSUSB
  if sheet.cell(row=rowNum,column=1).value == 'PSU Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send PS Query: " + MESSAGE
      try:
        data = P1.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close PSUSB
  if sheet.cell(row=rowNum,column=1).value == 'PSU Close':
      print rowNum,
      P1.close()
      print "*** GEN USB Disconnected ***"

  #*** Power Supply GEN: IEEE ****************************************************************
  #Open PSGPIB
  if sheet.cell(row=rowNum,column=1).value == 'PSG Open':
      try:
        P1 = rm.open_resource(GEN_IEEE)
      except visa.VisaIOError:
        print "*** GEN GPIB Not Found ***"
        e = 1
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      if e != 1:
        print "*** GEN GPIB Connected ***"
        ind = 0
        
  #Command PSGPIB
  if sheet.cell(row=rowNum,column=1).value == 'PSG Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      P1.write(MESSAGE)
      print " Send PS Command: " + MESSAGE
      sleep(delay)
      
  #Query PSGPIB
  if sheet.cell(row=rowNum,column=1).value == 'PSG Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send PS Query: " + MESSAGE
      try:
        data = P1.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close PSGPIB
  if sheet.cell(row=rowNum,column=1).value == 'PSG Close':
      print rowNum,
      P1.close()
      print "*** GEN GPIB Disconnected ***"
      
  #*** DMM1 ****************************************************************
  #Open DMM1
  if sheet.cell(row=rowNum,column=1).value == 'DMM1 Open':
      try:
        d1 = rm.open_resource("TCPIP0::" + DMM1_IP + "::inst0::INSTR")
      except visa.VisaIOError:
        print "*** DMM1 Not Found ***"
        e = 1
      #d1.timeout(3)
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      if e != 1:
        print "*** DMM1 Connected ***"
        ind = 0

  #Command DMM 1
  if sheet.cell(row=rowNum,column=1).value == 'DMM1 Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      d1.write(MESSAGE)
      print " Send DMM1 Command: " + MESSAGE
      sleep(delay)
      
  #Query DMM 1
  if sheet.cell(row=rowNum,column=1).value == 'DMM1 Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send DMM1 Query: " + MESSAGE
      try:
        data = d1.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close DMM1
  if sheet.cell(row=rowNum,column=1).value == 'DMM1 Close':
      print rowNum,
      d1.close()
      print "*** DMM1 Disconnected ***"


      
  #*** DMM2 ****************************************************************
  #Open DMM2
  if sheet.cell(row=rowNum,column=1).value == 'DMM2 Open':
      try:
        d2 = rm.open_resource("TCPIP0::" + DMM2_IP + "::inst0::INSTR")
      except visa.VisaIOError:
        print "*** DMM2 Not Found ***"
        e = 1
      #d1.timeout(3)
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      if e != 1:
        print "*** DMM2 Connected ***"
        ind = 0

  #Command DMM2
  if sheet.cell(row=rowNum,column=1).value == 'DMM2 Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      d2.write(MESSAGE)
      print " Send DMM2 Command: " + MESSAGE
      sleep(delay)
      
  #Query DMM2
  if sheet.cell(row=rowNum,column=1).value == 'DMM2 Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send DMM2 Query: " + MESSAGE
      try:
        data = d2.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close DMM2
  if sheet.cell(row=rowNum,column=1).value == 'DMM2 Close':
      print rowNum,
      d2.close()
      print "*** DMM2 Disconnected ***"

  #*** DMM3 ****************************************************************
  #Open DMM3
  if sheet.cell(row=rowNum,column=1).value == 'DMM3 Open':
      try:
        d3 = rm.open_resource("TCPIP0::" + DMM3_IP + "::inst0::INSTR")
      except visa.VisaIOError:
        print "*** DMM3 Not Found ***"
        e = 1
      #d1.timeout(3)
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      if e != 1:
        print "*** DMM3 Connected ***"
        ind = 0

  #Command DMM 3
  if sheet.cell(row=rowNum,column=1).value == 'DMM3 Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      d3.write(MESSAGE)
      print " Send DMM3 Command: " + MESSAGE
      sleep(delay)
      
  #Query DMM 3
  if sheet.cell(row=rowNum,column=1).value == 'DMM3 Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send DMM3 Query: " + MESSAGE
      try:
        data = d3.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close DMM3
  if sheet.cell(row=rowNum,column=1).value == 'DMM3 Close':
      print rowNum,
      d3.close()
      print "*** DMM3 Disconnected ***"

  #*** Chroma Load ****************************************************************
  #Open Load
  if sheet.cell(row=rowNum,column=1).value == 'Load Open':
      try:
        L1 = rm.open_resource("TCPIP0::" + LOAD_IP + "::inst0::INSTR")
      except visa.VisaIOError:
        print "*** Load Not Found ***"
        e = 1
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      if e!= 1:
        print "*** Load Connected ***"
        ind = 0

  #Command Load
  if sheet.cell(row=rowNum,column=1).value == 'Load Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      L1.write(MESSAGE)
      print " Send DMM2 Command: " + MESSAGE
      sleep(delay)
      
  #Query Load
  if sheet.cell(row=rowNum,column=1).value == 'Load Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send Load Query: " + MESSAGE
      try:
        data = L1.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close Load
  if sheet.cell(row=rowNum,column=1).value == 'Load Close':
      print rowNum,
      L1.close()
      print "*** Load Disconnected ***"

  #*** Yokogawa Power Meter ****************************************************************
  #Open DPM
  if sheet.cell(row=rowNum,column=1).value == 'PM Open':
      try:
        P1 = rm.open_resource(DPM_USB)
      except visa.VisaIOError:
        print "*** Power Meter Not Found ***"
        e = 1
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      if e != 1:
        print "*** Power Meter Connected ***"
        ind = 0

  #Command DPM
  if sheet.cell(row=rowNum,column=1).value == 'PM Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      P1.write(MESSAGE)
      print " Send PM Command: " + MESSAGE
      sleep(delay)
      
  #Query DPM
  if sheet.cell(row=rowNum,column=1).value == 'PM Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send PM Query: " + MESSAGE
      try:
        data = P1.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close DPM
  if sheet.cell(row=rowNum,column=1).value == 'PM Close':
      print rowNum,
      P1.close()
      print "*** Power Meter Disconnected ***"


  #*** Yokogawa Scope ****************************************************************
  #Open Scope
  if sheet.cell(row=rowNum,column=1).value == 'Scope Open':
      try:
        O1 = rm.open_resource(SCOPE_USB)
      except visa.VisaIOError:
        print "*** Oscilloscope Not Found ***"
        e = 1
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      if e != 1:
        print "*** Oscilloscope Connected ***"
        ind = 0

  #Command Scope
  if sheet.cell(row=rowNum,column=1).value == 'Scope Command':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*" 
      print rowNum,
      ind = 0
      O1.write(MESSAGE)
      print " Send Scope Command: " + MESSAGE
      sleep(delay)
      
  #Query Scope
  if sheet.cell(row=rowNum,column=1).value == 'Scope Query':
      MESSAGE = sheet.cell(row=rowNum,column=2).value + ";"
      if ind == 1:
        print "           "
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
      print rowNum,
      ind = 0
      print " Send Scope Query: " + MESSAGE
      try:
        data = O1.query(MESSAGE).strip()
      except visa.VisaIOError:
        data = "ERR: Time-out"
      sheet.cell(row=rowNum,column=3).value = data #Responses in 3rd column
      sleep(delay)
      dataC = data
      expectC = sheet.cell(row=rowNum,column=4).value
      print "     Response:    " + dataC.rstrip()
      
  #Close Scope
  if sheet.cell(row=rowNum,column=1).value == 'Scope Close':
      print rowNum,
      O1.close()
      print "*** Oscilloscope Disconnected ***"


      
  #**** GENERAL ***********************************************************
  #Pause
  if sheet.cell(row=rowNum,column=1).value == 'Pause':
    ind = 0
    print "         "
    pause_continue = raw_input("Press Enter to continue")
    print "           "
    print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
    
  #Comment
  if sheet.cell(row=rowNum,column=1).value == 'Comment':
    MESSAGE = sheet.cell(row=rowNum,column=2).value
    if ind == 0:
        print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
        print "     "
    ind = 1
    print MESSAGE

  #Wait
  if sheet.cell(row=rowNum,column=1).value == 'Wait':
      wait = sheet.cell(row=rowNum,column=2).value
      print "    Wait for: " + str(wait) + " second(s)"
      sleep(wait)

  #Loop
  if sheet.cell(row=rowNum,column=1).value == 'Loop':
      LOOP_AMOUNT = float(LOOP_AMOUNT)
      loopCount += 1
      if LONG_TERM_LOG == "1":      #Long Term Log Save
          sheet.cell(row=3,column=5).value = "Total Errors: " + str(errCount)
          sheet.cell(row=2,column=5).value = "Time-outs: " + str(timCount)
          sheet.cell(row=1,column=5).value = "Commands Sent: " + str(comCount)
          sheet.cell(row=rowNum,column=3).value = "Total Loops: " + str(loopCount)
          if prot == 1:
              wb.save(EXCEL_RESULTS)
          if prot == 2:
              wb.save(UDP_EXCEL_RESULTS)
          if prot == 3:
              wb.save(VISA_EXCEL_RESULTS)
      wait = sheet.cell(row=rowNum,column=2).value
      print "Loops: " + str(loopCount)
      if wait == None:
          wait = float(0)
          print "Total errors: " + str(errCount)
      else:
          print "Total errors: " + str(errCount)
          print "Looping in " + str(wait) + " second(s)..."
      sleep(wait)
      if loopCount < LOOP_AMOUNT:
          print "    *****Loop: Back to start*****" 
          rowNum = 4
      elif loopCount >= LOOP_AMOUNT:
          print "    *****Loop Ended*****"
          rowNum = MaxRow
      
    
  #Data Compare: Pass/Fail
  N = sheet.cell(row=rowNum,column=1).value
  if dataC == expectC and N is not None and rowNum > 4: #Pass: Green
   sheet.cell(row=rowNum,column=5).fill = PatternFill(fgColor="29E31F", fill_type = "solid")
   dataC = 0
   expectC = 0
  if dataC != expectC and expectC is not None or timFlag == 1: #Fail: Red
   sheet.cell(row=rowNum,column=5).fill = PatternFill(fgColor="FF3B33", fill_type = "solid")
   dataC = 0
   expectC = 0
   timFlag = 0
   errCount += 1
  elif expectC is None: #Response doesnt matter: Blue
   sheet.cell(row=rowNum,column=5).fill = PatternFill(fgColor="00C8FF", fill_type = "solid")
   dataC = 0
   expectC = 0
   
  #Increment Row
  rowNum +=1


# Print counters in Results
sheet.cell(row=3,column=5).value = "Total Errors: " + str(errCount)
sheet.cell(row=2,column=5).value = "Time-outs: " + str(timCount)
sheet.cell(row=1,column=5).value = "Commands Sent: " + str(comCount)

# Save Results File
wb.save(EXCEL_RESULTS)


print "#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*#*"
print "   "
print "End of List Reached"
print "Errors: " + str(errCount)

           
